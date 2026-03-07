import io
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="2C5F8A", end_color="2C5F8A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOTAL_FONT = Font(bold=True)


class ExportService:
    def generate_excel(self, data: Dict[str, Any]) -> bytes:
        wb = Workbook()

        self._create_resume_sheet(wb, data)
        self._create_main_oeuvre_sheet(wb, data)
        self._create_lignes_sheet(wb, data)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _create_resume_sheet(self, wb: Workbook, data: Dict[str, Any]):
        ws = wb.active
        ws.title = "Résumé chantier"

        rows = [
            ("Chantier", data.get("nom", "")),
            ("Date", data.get("date", "")),
            ("Métier", data.get("metier", "")),
            ("", ""),
            ("CA (Chiffre d'affaires)", data.get("ca", 0)),
            ("Coût matériaux", data.get("cout_materiaux", 0)),
            ("Coût main d'œuvre", data.get("cout_mo", 0)),
            ("Coût total", data.get("cout_total", 0)),
            ("", ""),
            ("Gain estimé", data.get("gain", 0)),
            ("Gain par jour", data.get("gain_par_jour", 0)),
            ("Rentabilité", data.get("badge_rentabilite", "")),
        ]

        for i, (label, value) in enumerate(rows, 1):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            cell = ws.cell(row=i, column=2, value=value)
            if isinstance(value, float) and label:
                cell.number_format = '#,##0.00 €'

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

    def _create_main_oeuvre_sheet(self, wb: Workbook, data: Dict[str, Any]):
        ws = wb.create_sheet("Main d'œuvre")
        headers = ["Nom", "Rôle", "Taux horaire (€/h)", "Temps prévu (h)", "Coût brut (€)"]
        self._write_headers(ws, headers)

        salaries = data.get("salaries", [])
        row = 2
        total_brut = 0.0
        for s in salaries:
            cout = s.get("taux_horaire", 0) * s.get("temps_prevu", 0)
            total_brut += cout
            ws.cell(row=row, column=1, value=s.get("nom", ""))
            ws.cell(row=row, column=2, value=s.get("role", ""))
            ws.cell(row=row, column=3, value=s.get("taux_horaire", 0)).number_format = '#,##0.00 €'
            ws.cell(row=row, column=4, value=s.get("temps_prevu", 0))
            ws.cell(row=row, column=5, value=cout).number_format = '#,##0.00 €'
            row += 1

        # Dirigeant
        dirigeant = data.get("dirigeant", {})
        if dirigeant:
            cout_dir = dirigeant.get("taux_horaire", 0) * dirigeant.get("temps_prevu", 0)
            total_brut += cout_dir
            ws.cell(row=row, column=1, value="Dirigeant")
            ws.cell(row=row, column=2, value="Dirigeant")
            ws.cell(row=row, column=3, value=dirigeant.get("taux_horaire", 0)).number_format = '#,##0.00 €'
            ws.cell(row=row, column=4, value=dirigeant.get("temps_prevu", 0))
            ws.cell(row=row, column=5, value=cout_dir).number_format = '#,##0.00 €'
            row += 1

        row += 1
        coeff = data.get("coefficient_charges", 1.45)
        cout_total_mo = total_brut * coeff
        ws.cell(row=row, column=4, value="Total brut").font = TOTAL_FONT
        ws.cell(row=row, column=5, value=total_brut).font = TOTAL_FONT
        ws.cell(row=row, column=5).number_format = '#,##0.00 €'
        row += 1
        ws.cell(row=row, column=4, value=f"Coefficient charges ({coeff})").font = TOTAL_FONT
        ws.cell(row=row, column=5, value=cout_total_mo).font = TOTAL_FONT
        ws.cell(row=row, column=5).number_format = '#,##0.00 €'

        self._auto_width(ws)

    def _create_lignes_sheet(self, wb: Workbook, data: Dict[str, Any]):
        ws = wb.create_sheet("Lignes devis")
        headers = [
            "Désignation", "Quantité", "Unité",
            "Prix facturé (€)", "Prix estimé (€)",
            "Écart (€)", "Type", "Confiance"
        ]
        self._write_headers(ws, headers)

        lignes = data.get("lignes", [])
        for i, ligne in enumerate(lignes, 2):
            prix_facture = ligne.get("prix_unitaire_facture", 0)
            prix_estime = ligne.get("prix_unitaire_estime") or 0
            ecart = prix_facture - prix_estime if prix_estime else None

            ws.cell(row=i, column=1, value=ligne.get("designation", ""))
            ws.cell(row=i, column=2, value=ligne.get("quantite", 0))
            ws.cell(row=i, column=3, value=ligne.get("unite", ""))
            ws.cell(row=i, column=4, value=prix_facture).number_format = '#,##0.00 €'
            ws.cell(row=i, column=5, value=prix_estime if prix_estime else "N/A")
            if prix_estime:
                ws.cell(row=i, column=5).number_format = '#,##0.00 €'
            ws.cell(row=i, column=6, value=ecart).number_format = '#,##0.00 €' if ecart else '#'
            ws.cell(row=i, column=7, value=ligne.get("type", ""))
            ws.cell(row=i, column=8, value=ligne.get("niveau_confiance") or "")

        self._auto_width(ws)

    def _write_headers(self, ws, headers: List[str]):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

    def _auto_width(self, ws):
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 50)
